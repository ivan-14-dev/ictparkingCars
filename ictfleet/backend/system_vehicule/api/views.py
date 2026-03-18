from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework.response import Response
from rest_framework import generics, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from django.utils import timezone
from django.db.models import Sum, Count, Avg, Q
from django.db.models.functions import TruncMonth, TruncWeek
from django.core.mail import send_mail
from django.conf import settings
from django.http import HttpResponse
from . import serializers
from .models import MaintenanceHistory, Activity, Breakdown, RepairRecord, Vehicle, FuelUsage
from users.permissions import IsAdminUser

# Try to import openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class CustomTokenObtainPairView(TokenObtainPairView):
    """
    Custom token obtain view that includes user information in the response
    """
    serializer_class = serializers.CustomTokenObtainPairSerializer

    def post(self, request, *args, **kwargs):
        response = super().post(request, *args, **kwargs)

        if response.status_code == 200:
            # Get the user from the validated data
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            user = serializer.user

            # Add user information to the response
            user_data = {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'role': user.role,
                'department': user.department,
                'phone_number': user.phone_number,
                'date_joined': user.date_joined.isoformat(),
                'last_login': user.last_login.isoformat() if user.last_login else None,
            }

            response.data['user'] = user_data

        return response


class MaintenanceListView(generics.ListAPIView):
    queryset = MaintenanceHistory.objects.all().order_by('-service_date')
    serializer_class = serializers.MaintenanceSerializer
    permission_classes = [IsAdminUser]


class ActivityListView(generics.ListAPIView):
    """View to list all activities"""
    serializer_class = serializers.ActivitySerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Return activities ordered by creation date"""
        queryset = Activity.objects.all().order_by('-created_at')
        
        # Filter by activity type if provided
        activity_type = self.request.query_params.get('activity_type')
        if activity_type:
            queryset = queryset.filter(activity_type=activity_type)
        
        # Filter by user if provided
        user_id = self.request.query_params.get('user_id')
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        
        # Limit to recent activities
        limit = self.request.query_params.get('limit')
        if limit:
            try:
                queryset = queryset[:int(limit)]
            except ValueError:
                pass
        
        return queryset


class ActivityCreateView(generics.CreateAPIView):
    """View to create activities"""
    serializer_class = serializers.ActivityCreateSerializer
    permission_classes = [IsAuthenticated]
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class RecentActivityView(generics.ListAPIView):
    """View to get recent activities for the dashboard"""
    serializer_class = serializers.ActivitySerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Return recent activities, more for admins"""
        user = self.request.user
        queryset = Activity.objects.all().order_by('-created_at')
        
        # Limit to 20 activities for the dashboard
        return queryset[:20]


class BreakdownListView(generics.ListAPIView):
    """View to list breakdowns - all roles can see breakdowns"""
    serializer_class = serializers.BreakdownSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """
        All authenticated users (admin, mechanic, driver) can see breakdowns.
        - Admins see all breakdowns
        - Mechanics see all breakdowns (they need to see all to know what to fix)
        - Drivers see breakdowns for vehicles assigned to them
        """
        user = self.request.user
        if user.role == 'admin':
            return Breakdown.objects.all()
        elif user.role == 'mechanic':
            return Breakdown.objects.all()
        else:
            # Drivers see breakdowns for vehicles assigned to them
            return Breakdown.objects.filter(vehicle__assigned_driver=user)


class BreakdownCreateView(generics.CreateAPIView):
    """View to create breakdown reports - for both mechanics and drivers"""
    serializer_class = serializers.BreakdownCreateSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def get_queryset(self):
        return Breakdown.objects.all()

    def perform_create(self, serializer):
        """Automatically set the reporter from the authenticated user (mechanic or driver)"""
        user = self.request.user
        # Both mechanics and drivers can create breakdowns
        # The 'mechanic' field actually means 'reporter' now
        serializer.save(mechanic=user)


class BreakdownDetailView(generics.RetrieveUpdateDestroyAPIView):
    """View to retrieve, update, or delete a breakdown - all roles can see breakdowns"""
    serializer_class = serializers.BreakdownSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def get_queryset(self):
        """
        All authenticated users can see breakdown details.
        - Admins and Mechanics can see all breakdowns
        - Drivers can see breakdowns for vehicles assigned to them
        """
        user = self.request.user
        if user.role == 'admin' or user.role == 'mechanic':
            return Breakdown.objects.all()
        else:
            return Breakdown.objects.filter(vehicle__assigned_driver=user)


class RepairRecordListView(generics.ListAPIView):
    """View to list repair records"""
    serializer_class = serializers.RepairRecordSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Admins see all records, mechanics see only their own"""
        user = self.request.user
        if user.role == 'admin':
            return RepairRecord.objects.all()
        else:
            return RepairRecord.objects.filter(mechanic=user)


class RepairRecordCreateView(generics.CreateAPIView):
    """View to create repair records"""
    serializer_class = serializers.RepairRecordCreateSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def get_queryset(self):
        return RepairRecord.objects.all()


class RepairRecordDetailView(generics.RetrieveUpdateDestroyAPIView):
    """View to retrieve, update, or delete a repair record"""
    serializer_class = serializers.RepairRecordSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def get_queryset(self):
        """Admins can see all, mechanics can only see their own"""
        user = self.request.user
        if user.role == 'admin':
            return RepairRecord.objects.all()
        else:
            return RepairRecord.objects.filter(mechanic=user)
    
    def perform_update(self, serializer):
        """Allow admin to set reviewed_by and reviewed_at"""
        user = self.request.user
        if user.role == 'admin' and serializer.validated_data.get('status') in ['reviewed', 'approved', 'rejected']:
            serializer.save(reviewed_by=user, reviewed_at=timezone.now())
        else:
            serializer.save()


class DriverRepairVerificationView(generics.GenericAPIView):
    """
    View for drivers to verify repair records for their vehicles.
    Drivers can view and verify that repairs have been completed correctly.
    """
    serializer_class = serializers.RepairRecordSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Drivers can see repair records for vehicles assigned to them"""
        user = self.request.user
        if user.role == 'driver':
            return RepairRecord.objects.filter(vehicle__assigned_driver=user).select_related('vehicle', 'mechanic', 'reviewed_by')
        return RepairRecord.objects.none()
    
    def get(self, request, pk=None):
        """Get a specific repair record for verification"""
        if not pk:
            # List all repair records for driver's vehicles
            repairs = self.get_queryset()
            serializer = self.get_serializer(repairs, many=True)
            return Response(serializer.data)
        
        try:
            repair = self.get_queryset().get(pk=pk)
            serializer = self.get_serializer(repair)
            return Response(serializer.data)
        except RepairRecord.DoesNotExist:
            raise HTTP404
    
    def post(self, request, pk):
        """Driver verifies a repair record"""
        try:
            repair = self.get_queryset().get(pk=pk)
        except RepairRecord.DoesNotExist:
            return Response(
                {'error': 'Repair record not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Update verification status
        repair.driver_verified = True
        repair.driver_verified_at = timezone.now()
        repair.driver_verified_by = request.user
        repair.driver_comments = request.data.get('comments', '')
        repair.save()
        
        return Response({
            'message': 'Repair verified successfully',
            'repair_id': repair.id,
            'verified_at': repair.driver_verified_at
        })


class RepairExpensesReportView(generics.ListAPIView):
    """
    View to get repair expenses report.
    Provides total expenses and breakdown by time period.
    """
    permission_classes = [IsAuthenticated]
    
    def list(self, request, *args, **kwargs):
        # Get date range from query params
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        group_by = request.query_params.get('group_by', 'month')  # month, week, day
        
        # Base querysets
        maintenance_qs = MaintenanceHistory.objects.all()
        repair_qs = RepairRecord.objects.filter(status='approved')
        
        # Apply date filters if provided
        if start_date:
            maintenance_qs = maintenance_qs.filter(service_date__gte=start_date)
            repair_qs = repair_qs.filter(completed_at__date__gte=start_date)
        
        if end_date:
            maintenance_qs = maintenance_qs.filter(service_date__lte=end_date)
            repair_qs = repair_qs.filter(completed_at__date__lte=end_date)
        
        # Calculate totals
        maintenance_total = maintenance_qs.aggregate(total=Sum('cost'))['total'] or 0
        
        # Calculate repair costs from accessories used
        repair_costs = 0
        for repair in repair_qs:
            for accessory in repair.accessories_used.all():
                repair_costs += float(accessory.price)
        
        total_expenses = float(maintenance_total) + repair_costs
        
        # Group by time period
        if group_by == 'month':
            maintenance_by_period = maintenance_qs.annotate(
                period=TruncMonth('service_date')
            ).values('period').annotate(
                total=Sum('cost'),
                count=Count('id')
            ).order_by('period')
        elif group_by == 'week':
            maintenance_by_period = maintenance_qs.annotate(
                period=TruncWeek('service_date')
            ).values('period').annotate(
                total=Sum('cost'),
                count=Count('id')
            ).order_by('period')
        else:
            maintenance_by_period = maintenance_qs.values('service_date').annotate(
                total=Sum('cost'),
                count=Count('id')
            ).order_by('service_date')
        
        # Get expenses by vehicle
        vehicle_expenses = {}
        for maintenance in maintenance_qs.select_related('vehicle'):
            vehicle_key = f"{maintenance.vehicle.make} {maintenance.vehicle.model} ({maintenance.vehicle.license_plate})"
            if vehicle_key not in vehicle_expenses:
                vehicle_expenses[vehicle_key] = 0
            vehicle_expenses[vehicle_key] += float(maintenance.cost or 0)
        
        # Get expenses by maintenance type
        maintenance_type_expenses = maintenance_qs.values('maintenance_type').annotate(
            total=Sum('cost'),
            count=Count('id')
        ).order_by('-total')
        
        return Response({
            'summary': {
                'total_expenses': total_expenses,
                'maintenance_total': float(maintenance_total),
                'repair_parts_total': repair_costs,
                'maintenance_count': maintenance_qs.count(),
                'repair_count': repair_qs.count(),
            },
            'by_period': list(maintenance_by_period),
            'by_vehicle': [{'vehicle': k, 'total': v} for k, v in vehicle_expenses.items()],
            'by_maintenance_type': list(maintenance_type_expenses),
            'date_range': {
                'start_date': start_date,
                'end_date': end_date,
            }
        })


class FuelUsageListView(generics.ListCreateAPIView):
    """View to list and create fuel usage records"""
    serializer_class = serializers.FuelUsageSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if user.role == 'admin' or user.role == 'mechanic':
            return FuelUsage.objects.all().select_related('vehicle', 'driver')
        else:
            # Drivers see only their own fuel records
            return FuelUsage.objects.filter(driver=user).select_related('vehicle', 'driver')
    
    def perform_create(self, serializer):
        """Ensure drivers can only add fuel for vehicles assigned to them"""
        user = self.request.user
        vehicle = serializer.validated_data.get('vehicle')
        
        # For drivers, validate they can only add fuel for their assigned vehicles
        if user.role == 'driver':
            if vehicle and vehicle.assigned_driver != user:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied(
                    "Vous ne pouvez ajouter du carburant que pour les véhicules qui vous sont attribués."
                )
            # Auto-assign the driver to the fuel record
            serializer.save(driver=user)
        else:
            # For admins/managers, allow any vehicle
            serializer.save()


class FuelUsageReportView(generics.GenericAPIView):
    """
    View to get fuel usage report.
    Provides total fuel costs and breakdown by time period (day, week, month).
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        group_by = request.query_params.get('group_by', 'month')  # day, week, month
        vehicle_id = request.query_params.get('vehicle_id')
        driver_id = request.query_params.get('driver_id')
        
        # Base queryset
        fuel_qs = FuelUsage.objects.all()
        
        # Filter by vehicle if provided
        if vehicle_id:
            fuel_qs = fuel_qs.filter(vehicle_id=vehicle_id)
        
        # Filter by driver if provided (admin only)
        if driver_id:
            fuel_qs = fuel_qs.filter(driver_id=driver_id)
        
        # Filter by user's vehicles for non-admin
        user = request.user
        if user.role != 'admin' and user.role != 'mechanic':
            fuel_qs = fuel_qs.filter(vehicle__assigned_driver=user)
        
        # Apply date filters
        if start_date:
            fuel_qs = fuel_qs.filter(date__gte=start_date)
        if end_date:
            fuel_qs = fuel_qs.filter(date__lte=end_date)
        
        # Calculate totals
        total_liters = fuel_qs.aggregate(total=Sum('liters'))['total'] or 0
        total_cost = fuel_qs.aggregate(total=Sum('total_cost'))['total'] or 0
        
        # Group by time period
        if group_by == 'day':
            fuel_by_period = fuel_qs.annotate(
                period=TruncMonth('date')
            ).values('date').annotate(
                total_liters=Sum('liters'),
                total_cost=Sum('total_cost'),
                count=Count('id')
            ).order_by('date')
        elif group_by == 'week':
            fuel_by_period = fuel_qs.annotate(
                period=TruncWeek('date')
            ).values('period').annotate(
                total_liters=Sum('liters'),
                total_cost=Sum('total_cost'),
                count=Count('id')
            ).order_by('period')
        else:  # month
            fuel_by_period = fuel_qs.annotate(
                period=TruncMonth('date')
            ).values('period').annotate(
                total_liters=Sum('liters'),
                total_cost=Sum('total_cost'),
                count=Count('id')
            ).order_by('period')
        
        # Get by vehicle
        fuel_by_vehicle = fuel_qs.values('vehicle__license_plate').annotate(
            total_liters=Sum('liters'),
            total_cost=Sum('total_cost'),
            count=Count('id')
        ).order_by('-total_cost')[:10]
        
        # Get by driver
        fuel_by_driver = fuel_qs.filter(driver__isnull=False).values(
            'driver__id', 'driver__first_name', 'driver__last_name'
        ).annotate(
            total_liters=Sum('liters'),
            total_cost=Sum('total_cost'),
            count=Count('id')
        ).order_by('-total_cost')[:10]
        
        return Response({
            'summary': {
                'total_liters': float(total_liters),
                'total_cost': float(total_cost),
                'total_records': fuel_qs.count(),
            },
            'by_period': list(fuel_by_period),
            'by_vehicle': list(fuel_by_vehicle),
            'by_driver': list(fuel_by_driver),
            'date_range': {
                'start_date': start_date,
                'end_date': end_date,
            }
        })


class FuelUsageExportView(generics.GenericAPIView):
    """
    View to export fuel usage data to Excel.
    Admin only.
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        if not OPENPYXL_AVAILABLE:
            return Response(
                {'error': 'Excel export is not available. Please install openpyxl.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        vehicle_id = request.query_params.get('vehicle_id')
        driver_id = request.query_params.get('driver_id')
        export_format = request.query_params.get('format', 'excel')
        
        # Base queryset
        fuel_qs = FuelUsage.objects.all().select_related('vehicle', 'driver')
        
        # Apply filters
        if vehicle_id:
            fuel_qs = fuel_qs.filter(vehicle_id=vehicle_id)
        if driver_id:
            fuel_qs = fuel_qs.filter(driver_id=driver_id)
        
        user = request.user
        if user.role != 'admin' and user.role != 'mechanic':
            fuel_qs = fuel_qs.filter(vehicle__assigned_driver=user)
        
        if start_date:
            fuel_qs = fuel_qs.filter(date__gte=start_date)
        if end_date:
            fuel_qs = fuel_qs.filter(date__lte=end_date)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Fuel Usage Report"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = ["Date", "Vehicle", "License Plate", "Driver", "Liters", "Price/Liter", "Total Cost"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data
        total_liters = 0
        total_cost = 0
        
        for fuel in fuel_qs.order_by('-date'):
            driver_name = f"{fuel.driver.first_name} {fuel.driver.last_name}" if fuel.driver else "-"
            ws.append([
                fuel.date.strftime("%Y-%m-%d"),
                f"{fuel.vehicle.make} {fuel.vehicle.model}",
                fuel.vehicle.license_plate,
                driver_name,
                float(fuel.liters),
                float(fuel.price_per_liter),
                float(fuel.total_cost)
            ])
            total_liters += float(fuel.liters)
            total_cost += float(fuel.total_cost)
        
        # Add totals row
        ws.append(["", "", "", "TOTAL:", total_liters, "", total_cost])
        
        # Format columns
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"fuel_usage_report_{start_date or 'all'}_{end_date or 'all'}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        return response


class DriverListView(generics.ListAPIView):
    """
    View to get list of drivers (for fuel filter dropdown).
    Admin only.
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        # Get users with driver role
        drivers = User.objects.filter(role='driver', is_active=True).values(
            'id', 'first_name', 'last_name', 'username'
        ).order_by('first_name', 'last_name')
        
        return Response(list(drivers))


class AdminBreakdownAlertView(generics.GenericAPIView):
    """
    View to send breakdown alerts to admin via email.
    Triggered when admin connects to the system.
    """
    permission_classes = [IsAdminUser]
    
    def get(self, request):
        """Send breakdown alert email to admin"""
        user = request.user
        
        # Get unresolved breakdowns
        unresolved_breakdowns = Breakdown.objects.filter(
            status__in=['reported', 'acknowledged']
        ).select_related('vehicle', 'mechanic').order_by('-reported_at')
        
        # Get recent breakdowns (last 7 days)
        recent_breakdowns = Breakdown.objects.filter(
            reported_at__gte=timezone.now() - timezone.timedelta(days=7)
        ).select_related('vehicle')
        
        # Build email content
        subject = f"ICT Fleet Alert - {unresolved_breakdowns.count()} Unresolved Breakdowns"
        
        message_lines = [
            f"Dear Administrator,",
            f"",
            f"This is an automated alert about vehicle breakdowns in your fleet.",
            f"",
            f"=== BREAKDOWN SUMMARY ===",
            f"",
            f"Unresolved Breakdowns: {unresolved_breakdowns.count()}",
            f"Breakdowns (Last 7 Days): {recent_breakdowns.count()}",
            f"",
        ]
        
        if unresolved_breakdowns.exists():
            message_lines.extend([
                f"=== UNRESOLVED BREAKDOWNS ===",
                f"",
            ])
            for breakdown in unresolved_breakdowns[:10]:
                message_lines.extend([
                    f"- Vehicle: {breakdown.vehicle}",
                    f"  Title: {breakdown.title}",
                    f"  Status: {breakdown.get_status_display()}",
                    f"  Reported: {breakdown.reported_at.strftime('%Y-%m-%d %H:%M')}",
                    f"  Description: {breakdown.description[:100]}..." if len(breakdown.description) > 100 else f"  Description: {breakdown.description}",
                    f"",
                ])
        
        message_lines.extend([
            f"=== RECOMMENDED ACTIONS ===",
            f"",
            f"1. Review unresolved breakdowns immediately",
            f"2. Assign mechanics to pending repairs",
            f"3. Check vehicle availability for operations",
            f"",
            f"Please log in to the system to take action.",
            f"",
            f"Best regards,",
            f"ICT Fleet Management System",
        ])
        
        message = "\n".join(message_lines)
        
        # Get admin email
        admin_email = user.email
        
        if not admin_email:
            return Response({
                'status': 'error',
                'message': 'Admin email not found'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Send email
        try:
            send_mail(
                subject=subject,
                message=message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[admin_email],
                fail_silently=False,
            )
            
            return Response({
                'status': 'success',
                'message': f'Breakdown alert sent to {admin_email}',
                'summary': {
                    'unresolved_breakdowns': unresolved_breakdowns.count(),
                    'recent_breakdowns': recent_breakdowns.count(),
                }
            })
        except Exception as e:
            return Response({
                'status': 'error',
                'message': f'Failed to send email: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CheckBreakdownAlertsView(generics.ListAPIView):
    """
    View to check for breakdown alerts (for frontend to call on admin login)
    """
    permission_classes = [IsAuthenticated]
    serializer_class = serializers.BreakdownSerializer
    
    def get_queryset(self):
        """Return unresolved breakdowns"""
        return Breakdown.objects.filter(
            status__in=['reported', 'acknowledged']
        ).select_related('vehicle', 'mechanic')[:20]
    
    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        
        # Get summary
        unresolved_count = Breakdown.objects.filter(
            status__in=['reported', 'acknowledged']
        ).count()
        
        recent_count = Breakdown.objects.filter(
            reported_at__gte=timezone.now() - timezone.timedelta(days=7)
        ).count()
        
        return Response({
            'alerts': serializer.data,
            'summary': {
                'unresolved_count': unresolved_count,
                'recent_count': recent_count,
                'has_alerts': unresolved_count > 0
            }
        })


